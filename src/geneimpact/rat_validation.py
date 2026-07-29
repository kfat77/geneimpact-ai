"""Pinned in-vivo rat guide-activity transfer benchmark."""

from __future__ import annotations

from dataclasses import dataclass
from hashlib import sha256
import math
from pathlib import Path
import re
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook


RAT_TRANSFER_REFERENCE = "https://doi.org/10.1038/s41592-018-0011-5"
RAT_TRANSFER_TABLE1_URL = (
    "https://static-content.springer.com/esm/art%3A10.1038%2F"
    "s41592-018-0011-5/MediaObjects/41592_2018_11_MOESM3_ESM.xlsx"
)
RAT_TRANSFER_TABLE5_URL = (
    "https://static-content.springer.com/esm/art%3A10.1038%2F"
    "s41592-018-0011-5/MediaObjects/41592_2018_11_MOESM7_ESM.xlsx"
)
RAT_TRANSFER_TABLE1_SHA256 = (
    "1efe7e4bb49a3feb9b5757b3c42989d0e95cd1da621ae4fd12c8f0aa3f10fdc8"
)
RAT_TRANSFER_TABLE5_SHA256 = (
    "7dfd554af5b9677723970a892c39ac2529a09894cb178f3252121fd18bd2e0c8"
)
_DESIGN_SEQUENCE_PATTERN = re.compile(r"^[ACGT]{19,20}$")
_ACTUAL_GUIDE_SEQUENCE_PATTERN = re.compile(r"^[ACGT]{20,21}$")
_SHA256_PATTERN = re.compile(r"^[0-9a-f]{64}$")
_MAX_WORKBOOK_BYTES = 20_000_000
_ALLOWED_SCORE_SEMANTICS = {"ranking_score", "probability"}
_ALLOWED_SEQUENCE_BASES = {"design_sequence", "actual_guide_sequence"}
_ALLOWED_OVERLAP_STATUS = {
    "declared_no_overlap",
    "unknown",
    "overlap_detected",
}


@dataclass(frozen=True)
class RatGuideActivitySource:
    source_id: str
    article_reference: str
    table1_url: str
    table5_url: str
    table1_sha256: str
    table5_sha256: str
    source_genome_build: str
    expected_rat_guide_count: int
    expected_label_count: int
    target_to_guide: tuple[tuple[str, str], ...]
    excluded_targets: tuple[str, ...]
    expected_total_observations: int | None = None
    license_note: str = (
        "Springer Nature supplementary material; source workbooks are not "
        "redistributed by GeneImpact AI."
    )


@dataclass(frozen=True)
class RatGuideTransferRecord:
    target: str
    prediction_input_sequence_sha256: str
    prediction_input_sequence_length: int
    predicted_score: float
    observed_mean_on_target_efficiency: float
    animal_or_embryo_count: int


@dataclass(frozen=True)
class RatGuideTransferMetrics:
    pearson_r: float
    pearson_ci95_lower: float
    pearson_ci95_upper: float
    spearman_rho: float
    mean_absolute_error: float | None
    root_mean_squared_error: float | None


@dataclass(frozen=True)
class RatGuideTransferReport:
    predictor: str
    predictor_version: str
    species_profile: str
    evaluation_status: str
    use: str
    source_id: str
    source_reference: str
    source_table1_url: str
    source_table5_url: str
    source_license_note: str
    source_genome_build: str
    current_registered_genome_build: str
    table1_sha256: str
    table5_sha256: str
    source_verification: str
    source_rat_guide_count: int
    source_label_count: int
    guide_count: int
    animal_or_embryo_observation_count: int
    excluded_ambiguous_guide_count: int
    excluded_targets: tuple[str, ...]
    score_semantics: str
    sequence_basis: str
    training_overlap_status: str
    training_overlap_evidence_reference: str
    independence_verified: bool
    independence_interpretation: str
    metrics: RatGuideTransferMetrics
    records: tuple[RatGuideTransferRecord, ...]
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _RatGuideSequences:
    design_sequence: str
    actual_guide_sequence: str


RAT_ANDERSON_2018_SOURCE = RatGuideActivitySource(
    source_id="anderson-2018-rat-in-vivo-guide-activity",
    article_reference=RAT_TRANSFER_REFERENCE,
    table1_url=RAT_TRANSFER_TABLE1_URL,
    table5_url=RAT_TRANSFER_TABLE5_URL,
    table1_sha256=RAT_TRANSFER_TABLE1_SHA256,
    table5_sha256=RAT_TRANSFER_TABLE5_SHA256,
    source_genome_build="rn5",
    expected_rat_guide_count=25,
    expected_label_count=16,
    target_to_guide=(
        ("rat Il13_ON", "rIl13_gRNA"),
        ("rat Map3k14_ON", "rMap3k14_gRNA"),
        ("rat Trpa1_ON", "rTrpa1_1_gRNA"),
        ("rat Usp30_ON", "rUsp30_1_gRNA"),
        ("rEsr1_sgRNA1_ON", "rEsr1_1_gRNA"),
        ("rEsr1_sgRNA2_ON", "rEsr1_2_gRNA"),
        ("rIL33_ON", "rIl33_gRNA"),
        ("rJag1_sgRNA1_ON", "rJag1_1_gRNA"),
        ("rJag1_sgRNA2_ON", "rJag1_2_gRNA"),
        ("rMap4k1_ON", "rMap4k1_gRNA"),
        ("rRipk1_ON", "rRipk1_gRNA"),
        ("rRorc_sgRNA1_ON", "rRorc_1_gRNA"),
        ("rRorc_sgRNA2_ON", "rRorc_2_gRNA"),
        ("rRorc_sgRNA3_ON", "rRorc_3_gRNA"),
    ),
    excluded_targets=(
        "rUsp30_sgRNA1_ON",
        "rUsp30_sgRNA2_ON",
    ),
    expected_total_observations=204,
)


def prepare_rat_guide_transfer_template(
    table1_path: Path,
    table5_path: Path,
    *,
    source: RatGuideActivitySource = RAT_ANDERSON_2018_SOURCE,
) -> dict[str, Any]:
    """Build a sequence-redacted prediction template from pinned workbooks."""
    table1_digest = _verified_digest(
        table1_path,
        source.table1_sha256,
        "Supplementary Table 1",
    )
    table5_digest = _verified_digest(
        table5_path,
        source.table5_sha256,
        "Supplementary Table 5",
    )
    guide_sequences, _ = _read_guides(table1_path, source)
    _read_labels(table5_path, source)
    return {
        "schema_version": "geneimpact.rat_guide_transfer_predictions.v1",
        "source": {
            "source_id": source.source_id,
            "reference": source.article_reference,
            "table1_sha256": table1_digest,
            "table5_sha256": table5_digest,
            "source_genome_build": source.source_genome_build,
            "license_note": source.license_note,
        },
        "prediction": {
            "name": "REPLACE_WITH_PREDICTOR_NAME",
            "version": "REPLACE_WITH_VERSION_OR_COMMIT",
            "score_direction": "higher_is_more_active",
            "score_semantics": "ranking_score",
            "sequence_basis": (
                "REPLACE_WITH_design_sequence_OR_actual_guide_sequence"
            ),
            "training_overlap_status": "unknown",
            "evidence_reference": "REPLACE_WITH_MODEL_OR_RUN_REFERENCE",
        },
        "records": [
            {
                "target": target,
                "design_sequence_sha256": sha256(
                    guide_sequences[guide].design_sequence.encode("ascii")
                ).hexdigest(),
                "design_sequence_length": len(
                    guide_sequences[guide].design_sequence
                ),
                "actual_guide_sequence_sha256": sha256(
                    guide_sequences[guide].actual_guide_sequence.encode("ascii")
                ).hexdigest(),
                "actual_guide_sequence_length": len(
                    guide_sequences[guide].actual_guide_sequence
                ),
                "predicted_score": None,
            }
            for target, guide in source.target_to_guide
        ],
        "instructions": (
            "Replace prediction metadata, declare whether the model consumed "
            "the design spacer or the 5'-G actual guide, choose ranking_score "
            "or probability, audit training overlap, and fill every "
            "predicted_score. Do not alter target names, lengths, or hashes."
        ),
    }


def evaluate_rat_guide_transfer(
    table1_path: Path,
    table5_path: Path,
    predictions: Mapping[str, Any],
    *,
    source: RatGuideActivitySource = RAT_ANDERSON_2018_SOURCE,
) -> RatGuideTransferReport:
    """Evaluate one external predictor on a pinned rat in-vivo guide set."""
    table1_digest = _verified_digest(
        table1_path,
        source.table1_sha256,
        "Supplementary Table 1",
    )
    table5_digest = _verified_digest(
        table5_path,
        source.table5_sha256,
        "Supplementary Table 5",
    )
    guide_sequences, source_rat_guide_count = _read_guides(table1_path, source)
    labels, source_label_count = _read_labels(table5_path, source)
    prediction_metadata, prediction_scores = _validate_predictions(
        predictions,
        source,
        guide_sequences,
    )

    records = tuple(
        RatGuideTransferRecord(
            target=target,
            prediction_input_sequence_sha256=sha256(
                getattr(
                    guide_sequences[guide],
                    prediction_metadata["sequence_basis"],
                ).encode("ascii")
            ).hexdigest(),
            prediction_input_sequence_length=len(
                getattr(
                    guide_sequences[guide],
                    prediction_metadata["sequence_basis"],
                )
            ),
            predicted_score=prediction_scores[target],
            observed_mean_on_target_efficiency=labels[target][0],
            animal_or_embryo_count=labels[target][1],
        )
        for target, guide in source.target_to_guide
    )
    predicted = [record.predicted_score for record in records]
    observed = [
        record.observed_mean_on_target_efficiency for record in records
    ]
    pearson = _pearson(predicted, observed)
    ci_lower, ci_upper = _pearson_ci95(pearson, len(records))
    spearman = _pearson(_average_ranks(predicted), _average_ranks(observed))
    probability_scores = prediction_metadata["score_semantics"] == "probability"
    mae = (
        sum(abs(a - b) for a, b in zip(predicted, observed, strict=True))
        / len(records)
        if probability_scores
        else None
    )
    rmse = (
        math.sqrt(
            sum(
                (a - b) ** 2
                for a, b in zip(predicted, observed, strict=True)
            )
            / len(records)
        )
        if probability_scores
        else None
    )
    overlap_status = prediction_metadata["training_overlap_status"]
    return RatGuideTransferReport(
        predictor=prediction_metadata["name"],
        predictor_version=prediction_metadata["version"],
        species_profile="rat",
        evaluation_status="retrospective_external_transfer_benchmark",
        use="external_transfer_ranking_benchmark_only",
        source_id=source.source_id,
        source_reference=source.article_reference,
        source_table1_url=source.table1_url,
        source_table5_url=source.table5_url,
        source_license_note=source.license_note,
        source_genome_build=source.source_genome_build,
        current_registered_genome_build="GRCr8",
        table1_sha256=table1_digest,
        table5_sha256=table5_digest,
        source_verification="pinned_workbooks_verified",
        source_rat_guide_count=source_rat_guide_count,
        source_label_count=source_label_count,
        guide_count=len(records),
        animal_or_embryo_observation_count=sum(
            record.animal_or_embryo_count for record in records
        ),
        excluded_ambiguous_guide_count=len(source.excluded_targets),
        excluded_targets=source.excluded_targets,
        score_semantics=prediction_metadata["score_semantics"],
        sequence_basis=prediction_metadata["sequence_basis"],
        training_overlap_status=overlap_status,
        training_overlap_evidence_reference=prediction_metadata[
            "evidence_reference"
        ],
        independence_verified=False,
        independence_interpretation=(
            "Training-overlap status is submitter-declared and has not been "
            "reproduced by GeneImpact AI."
        ),
        metrics=RatGuideTransferMetrics(
            pearson_r=pearson,
            pearson_ci95_lower=ci_lower,
            pearson_ci95_upper=ci_upper,
            spearman_rho=spearman,
            mean_absolute_error=mae,
            root_mean_squared_error=rmse,
        ),
        records=records,
        warnings=(
            "Only 14 uniquely mapped guides are evaluated; two Usp30 labels "
            "are excluded because the source-table guide mapping is ambiguous.",
            "The source labels are highly selected and guide-level sample size "
            "is too small for model training or probability calibration.",
            "The source coordinates use rn5, not the registered GRCr8 assembly; "
            "sequence hashes bind comparisons without implying coordinate lift-over.",
            "Animal or embryo counts are not independent guide-level replicates.",
            "Training-overlap status is self-declared; this report does not "
            "establish an independently verified test set.",
            "Correlation does not establish phenotype, off-target safety, animal "
            "welfare, or performance for a new strain, locus, delivery method, "
            "laboratory, or edit class.",
        ),
    )


def _verified_digest(path: Path, expected: str, label: str) -> str:
    if not _SHA256_PATTERN.fullmatch(expected):
        raise ValueError(f"{label} expected SHA-256 is invalid.")
    size = path.stat().st_size
    if not 1 <= size <= _MAX_WORKBOOK_BYTES:
        raise ValueError(f"{label} must be between 1 byte and 20 MB.")
    digest = sha256(path.read_bytes()).hexdigest()
    if digest != expected:
        raise ValueError(f"{label} SHA-256 does not match the pinned source.")
    return digest


def _read_guides(
    path: Path,
    source: RatGuideActivitySource,
) -> tuple[dict[str, _RatGuideSequences], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "gRNA ON_OFF Target List" not in workbook.sheetnames:
            raise ValueError(
                "Supplementary Table 1 is missing gRNA ON_OFF Target List."
            )
        sheet = workbook["gRNA ON_OFF Target List"]
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows)
        index = _header_index(
            headers,
            (
                "Name",
                "5'+1nt Target Sequence\n(Actual)",
                "Target Sequence\n(Design)",
                "Genome",
            ),
            "Supplementary Table 1",
        )
        rat_guides: dict[str, _RatGuideSequences] = {}
        for row in rows:
            name = row[index["Name"]]
            genome = row[index["Genome"]]
            if (
                genome != source.source_genome_build
                or not isinstance(name, str)
                or not name.endswith("_gRNA")
            ):
                continue
            design_sequence = str(
                row[index["Target Sequence\n(Design)"]]
            ).upper()
            actual_guide_sequence = str(
                row[index["5'+1nt Target Sequence\n(Actual)"]]
            ).upper()
            if not _DESIGN_SEQUENCE_PATTERN.fullmatch(design_sequence):
                raise ValueError(
                    f"invalid 19- or 20-nt design sequence for source guide {name}."
                )
            if not _ACTUAL_GUIDE_SEQUENCE_PATTERN.fullmatch(
                actual_guide_sequence
            ):
                raise ValueError(
                    f"invalid 20- or 21-nt actual sequence for source guide {name}."
                )
            if actual_guide_sequence != f"G{design_sequence}":
                raise ValueError(
                    f"actual sequence for source guide {name} is not 5'-G "
                    "plus the design sequence."
                )
            if name in rat_guides:
                raise ValueError(f"duplicate source guide {name}.")
            rat_guides[name] = _RatGuideSequences(
                design_sequence=design_sequence,
                actual_guide_sequence=actual_guide_sequence,
            )
        if len(rat_guides) != source.expected_rat_guide_count:
            raise ValueError(
                "Supplementary Table 1 rat guide count does not match the "
                "qualified source profile."
            )
        required = {guide for _, guide in source.target_to_guide}
        missing = required - rat_guides.keys()
        if missing:
            raise ValueError(
                "Supplementary Table 1 is missing qualified guides: "
                + ", ".join(sorted(missing))
            )
        return rat_guides, len(rat_guides)
    finally:
        workbook.close()


def _read_labels(
    path: Path,
    source: RatGuideActivitySource,
) -> tuple[dict[str, tuple[float, int]], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    expected_targets = {
        target for target, _ in source.target_to_guide
    } | set(source.excluded_targets)
    labels: dict[str, tuple[float, int]] = {}
    try:
        for sheet_name in (
            "Suppl.Fig.8a OT-positive",
            "Suppl.Fig.8a OT-negative",
        ):
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Supplementary Table 5 is missing {sheet_name}."
                )
            rows = workbook[sheet_name].iter_rows(values_only=True)
            headers = next(rows)
            index = _header_index(
                headers,
                (
                    "Target",
                    "Animal number (n)",
                ),
                "Supplementary Table 5",
            )
            efficiency_index = _single_header_alias(
                headers,
                (
                    "Mean ON-target efficiency",
                    "Mean ON target efficiency",
                ),
                "Supplementary Table 5",
            )
            for row in rows:
                target = row[index["Target"]]
                if target not in expected_targets:
                    continue
                if target in labels:
                    raise ValueError(
                        f"duplicate Supplementary Table 5 target {target}."
                    )
                efficiency = _fraction(
                    row[efficiency_index],
                    f"observed efficiency for {target}",
                )
                count = row[index["Animal number (n)"]]
                if isinstance(count, bool) or not isinstance(count, int) or count < 1:
                    raise ValueError(
                        f"animal number for {target} must be a positive integer."
                    )
                labels[str(target)] = (efficiency, count)
        if len(labels) != source.expected_label_count:
            raise ValueError(
                "Supplementary Table 5 qualified label count does not match "
                "the source profile."
            )
        if source.expected_total_observations is not None and sum(
            count for _, count in labels.values()
        ) != source.expected_total_observations:
            raise ValueError(
                "Supplementary Table 5 observation count does not match the "
                "source profile."
            )
        return labels, len(labels)
    finally:
        workbook.close()


def _validate_predictions(
    document: Mapping[str, Any],
    source: RatGuideActivitySource,
    guide_sequences: Mapping[str, _RatGuideSequences],
) -> tuple[dict[str, str], dict[str, float]]:
    metadata = document.get("prediction")
    raw_records = document.get("records")
    if not isinstance(metadata, Mapping):
        raise ValueError("prediction must be an object.")
    if not isinstance(raw_records, Sequence) or isinstance(
        raw_records, (str, bytes)
    ):
        raise ValueError("records must be a list.")
    values = {
        key: str(metadata.get(key, "")).strip()
        for key in (
            "name",
            "version",
            "score_direction",
            "score_semantics",
            "sequence_basis",
            "training_overlap_status",
            "evidence_reference",
        )
    }
    if not values["name"] or not values["version"] or not values["evidence_reference"]:
        raise ValueError(
            "prediction name, version, and evidence_reference are required."
        )
    if values["score_direction"] != "higher_is_more_active":
        raise ValueError("score_direction must be higher_is_more_active.")
    if values["score_semantics"] not in _ALLOWED_SCORE_SEMANTICS:
        raise ValueError(
            "score_semantics must be ranking_score or probability."
        )
    if values["sequence_basis"] not in _ALLOWED_SEQUENCE_BASES:
        raise ValueError(
            "sequence_basis must be design_sequence or actual_guide_sequence."
        )
    if values["training_overlap_status"] not in _ALLOWED_OVERLAP_STATUS:
        raise ValueError(
            "training_overlap_status must be declared_no_overlap, unknown, "
            "or overlap_detected."
        )

    expected = dict(source.target_to_guide)
    if len(raw_records) != len(expected):
        raise ValueError(
            f"records must contain exactly {len(expected)} qualified guides."
        )
    scores: dict[str, float] = {}
    for index, raw in enumerate(raw_records, start=1):
        if not isinstance(raw, Mapping):
            raise ValueError(f"record {index} must be an object.")
        target = str(raw.get("target", "")).strip()
        if target not in expected or target in scores:
            raise ValueError(f"record {index} has an unknown or duplicate target.")
        guide = expected[target]
        sequences = guide_sequences[guide]
        for field_name in _ALLOWED_SEQUENCE_BASES:
            sequence = getattr(sequences, field_name)
            if raw.get(f"{field_name}_sha256") != sha256(
                sequence.encode("ascii")
            ).hexdigest():
                raise ValueError(
                    f"record {index} {field_name}_sha256 does not match the source."
                )
            if raw.get(f"{field_name}_length") != len(sequence):
                raise ValueError(
                    f"record {index} {field_name}_length does not match the source."
                )
        score = _number(raw.get("predicted_score"), f"record {index} predicted_score")
        if values["score_semantics"] == "probability" and not 0 <= score <= 1:
            raise ValueError(
                f"record {index} probability score must be between 0 and 1."
            )
        scores[target] = score
    return values, scores


def _header_index(
    headers: Sequence[Any],
    required: Sequence[str],
    label: str,
) -> dict[str, int]:
    positions = {str(value): index for index, value in enumerate(headers)}
    missing = [name for name in required if name not in positions]
    if missing:
        raise ValueError(f"{label} is missing columns: {', '.join(missing)}")
    return {name: positions[name] for name in required}


def _single_header_alias(
    headers: Sequence[Any],
    allowed: Sequence[str],
    label: str,
) -> int:
    positions = {str(value): index for index, value in enumerate(headers)}
    matches = [name for name in allowed if name in positions]
    if len(matches) != 1:
        raise ValueError(
            f"{label} must contain exactly one of: {', '.join(allowed)}"
        )
    return positions[matches[0]]


def _number(value: Any, label: str) -> float:
    if isinstance(value, bool):
        raise ValueError(f"{label} must be numeric.")
    try:
        number = float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{label} must be numeric.") from error
    if not math.isfinite(number):
        raise ValueError(f"{label} must be finite.")
    return number


def _fraction(value: Any, label: str) -> float:
    number = _number(value, label)
    if not 0 <= number <= 1:
        raise ValueError(f"{label} must be between 0 and 1.")
    return number


def _pearson(x: Sequence[float], y: Sequence[float]) -> float:
    mean_x = sum(x) / len(x)
    mean_y = sum(y) / len(y)
    centered = [(a - mean_x, b - mean_y) for a, b in zip(x, y, strict=True)]
    denominator = math.sqrt(
        sum(a * a for a, _ in centered)
        * sum(b * b for _, b in centered)
    )
    if denominator == 0:
        raise ValueError("correlation is undefined for a constant series.")
    return sum(a * b for a, b in centered) / denominator


def _average_ranks(values: Sequence[float]) -> list[float]:
    ordered = sorted(enumerate(values), key=lambda item: item[1])
    ranks = [0.0] * len(values)
    start = 0
    while start < len(ordered):
        end = start + 1
        while end < len(ordered) and ordered[end][1] == ordered[start][1]:
            end += 1
        average_rank = (start + 1 + end) / 2
        for ordered_index in range(start, end):
            ranks[ordered[ordered_index][0]] = average_rank
        start = end
    return ranks


def _pearson_ci95(correlation: float, sample_size: int) -> tuple[float, float]:
    bounded = min(max(correlation, -0.999999999999), 0.999999999999)
    fisher_z = math.atanh(bounded)
    margin = 1.959963984540054 / math.sqrt(sample_size - 3)
    return math.tanh(fisher_z - margin), math.tanh(fisher_z + margin)
