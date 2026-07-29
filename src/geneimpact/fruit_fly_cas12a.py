"""Pinned array-level in-vivo Cas12a evidence for fruit fly."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from hashlib import sha256
from io import BytesIO, StringIO
from itertools import islice
from pathlib import Path
import re

from openpyxl import load_workbook


FRUIT_FLY_CAS12A_REFERENCE = (
    "https://doi.org/10.1038/s41467-026-68434-z"
)
FRUIT_FLY_CAS12A_LIBRARY_URL = (
    "https://static-content.springer.com/esm/art%3A10.1038%2F"
    "s41467-026-68434-z/MediaObjects/"
    "41467_2026_68434_MOESM3_ESM.csv"
)
FRUIT_FLY_CAS12A_GENOTYPES_URL = (
    "https://static-content.springer.com/esm/art%3A10.1038%2F"
    "s41467-026-68434-z/MediaObjects/"
    "41467_2026_68434_MOESM5_ESM.xlsx"
)
FRUIT_FLY_CAS12A_SOURCE_DATA_URL = (
    "https://static-content.springer.com/esm/art%3A10.1038%2F"
    "s41467-026-68434-z/MediaObjects/"
    "41467_2026_68434_MOESM9_ESM.xlsx"
)
FRUIT_FLY_CAS12A_LIBRARY_SHA256 = (
    "c8817630434131182784a2f46572f78da718d267a8eb132e51ff81bab696b78d"
)
FRUIT_FLY_CAS12A_GENOTYPES_SHA256 = (
    "1f084d138cb55db1ea1ad94c0905f4fad19e0c0c5c4f2d39cac97317a083d9cf"
)
FRUIT_FLY_CAS12A_SOURCE_DATA_SHA256 = (
    "ec3e73f35dd0a9b31fc17e5422eba185840f72ab7ba7398489c31b75efb586f1"
)

_MAX_SOURCE_BYTES = 20_000_000
_MAX_SOURCE_ROWS = 20_000
_SHA256_PATTERN = re.compile(r"^[0-9a-f]{64}$")
_LINE_ID_PATTERN = re.compile(r"^HD12ACFD([0-9]{1,4})(L?)$")
_GUIDE_PATTERN = re.compile(r"^[ACGT]{23}$")
_GENE_ID_PATTERN = re.compile(r"^FBgn[0-9]{7}$")
_LIBRARY_HEADERS = (
    "Line_ID",
    "sgRNA_1",
    "sgRNA_2",
    "sgRNA_3",
    "sgRNA_4",
    "primary_gene",
    "primary_gene_symbol",
)
_FIG5I_HEADERS = (
    "Line_ID",
    "CIMR_score",
    "Comment",
    "Disc_size",
    "Date",
)
_FIG5J_HEADERS = (
    "sgRNA_Line",
    "LOH_0",
    "LOH_1",
    "LOH_2",
    "LOH_3",
    "comment",
)


@dataclass(frozen=True)
class FruitFlyCas12aSource:
    source_id: str
    article_reference: str
    library_url: str
    genotypes_url: str
    source_data_url: str
    library_sha256: str
    genotypes_sha256: str
    source_data_sha256: str
    expected_library_array_count: int
    expected_library_guide_count: int
    expected_fig5i_row_count: int
    expected_fig5i_numeric_array_observation_count: int
    expected_fig5i_missing_array_observation_count: int
    expected_fig5i_control_observation_count: int
    expected_fig5i_unique_array_count: int
    expected_fig5j_row_count: int
    expected_fig5j_scored_array_count: int
    expected_fig5j_disc_count: int
    article_reported_2l_line_count: int | None = None
    article_reported_2l_guide_count: int | None = None
    article_reported_2r_line_count: int | None = None
    article_reported_2r_guide_count: int | None = None
    article_reported_on_target_active_arrays: int | None = None
    article_reported_on_target_tested_arrays: int | None = None
    license_note: str = (
        "Article and source data are licensed CC BY 4.0; GeneImpact AI "
        "does not bundle the publisher files."
    )


@dataclass(frozen=True)
class FruitFlyCas12aEvidenceAudit:
    species_profile: str
    nuclease: str
    reagent_level: str
    evidence_task: str
    benchmark_status: str
    source_id: str
    source_reference: str
    source_license_note: str
    library_url: str
    genotypes_url: str
    source_data_url: str
    library_sha256: str
    genotypes_sha256: str
    source_data_sha256: str
    source_verification: str
    library_array_count: int
    library_guide_count: int
    three_guide_array_count: int
    four_guide_array_count: int
    fig5i_row_count: int
    fig5i_numeric_array_observation_count: int
    fig5i_missing_array_observation_count: int
    fig5i_control_observation_count: int
    fig5i_unique_array_count: int
    fig5j_row_count: int
    fig5j_scored_array_count: int
    fig5j_disc_count: int
    article_reported_2l_line_count: int | None
    article_reported_2l_guide_count: int | None
    article_reported_2r_line_count: int | None
    article_reported_2r_guide_count: int | None
    article_reported_on_target_active_arrays: int | None
    article_reported_on_target_tested_arrays: int | None
    article_aggregate_reconstruction_status: str
    interval_manifest_status: str
    predictive_adapter_available: bool
    discrimination_metrics_status: str
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class FruitFlyCas12aArrayEvidence:
    species_profile: str
    nuclease: str
    source_id: str
    source_line_id: str
    primary_gene_id: str
    primary_gene_symbol: str | None
    component_guide_count: int
    component_sequence_sha256: tuple[str, ...]
    fig5i_numeric_observation_count: int
    fig5i_missing_observation_count: int
    fig5i_score_counts_0_to_3: tuple[int, int, int, int]
    fig5i_mean_score: float | None
    fig5j_disc_count: int
    fig5j_score_counts_0_to_3: tuple[int, int, int, int] | None
    fig5j_mean_score: float | None
    screen_membership: str
    interval_relationship: str
    interpretation: str
    calibration_status: str
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _LibraryArray:
    source_line_id: str
    primary_gene_id: str
    primary_gene_symbol: str | None
    guides: tuple[str, ...]


@dataclass(frozen=True)
class _ParsedEvidence:
    arrays: dict[str, _LibraryArray]
    fig5i_scores: dict[str, tuple[int, ...]]
    fig5i_missing: dict[str, int]
    fig5j_counts: dict[str, tuple[int, int, int, int] | None]
    fig5i_row_count: int
    fig5i_numeric_array_observation_count: int
    fig5i_missing_array_observation_count: int
    fig5i_control_observation_count: int
    fig5i_unique_array_count: int
    fig5j_row_count: int
    fig5j_scored_array_count: int
    fig5j_disc_count: int


PORT_2026_CAS12A_SOURCE = FruitFlyCas12aSource(
    source_id="port-2026-dmel-cas12a-array-loh",
    article_reference=FRUIT_FLY_CAS12A_REFERENCE,
    library_url=FRUIT_FLY_CAS12A_LIBRARY_URL,
    genotypes_url=FRUIT_FLY_CAS12A_GENOTYPES_URL,
    source_data_url=FRUIT_FLY_CAS12A_SOURCE_DATA_URL,
    library_sha256=FRUIT_FLY_CAS12A_LIBRARY_SHA256,
    genotypes_sha256=FRUIT_FLY_CAS12A_GENOTYPES_SHA256,
    source_data_sha256=FRUIT_FLY_CAS12A_SOURCE_DATA_SHA256,
    expected_library_array_count=845,
    expected_library_guide_count=3373,
    expected_fig5i_row_count=8197,
    expected_fig5i_numeric_array_observation_count=8041,
    expected_fig5i_missing_array_observation_count=142,
    expected_fig5i_control_observation_count=14,
    expected_fig5i_unique_array_count=536,
    expected_fig5j_row_count=600,
    expected_fig5j_scored_array_count=490,
    expected_fig5j_disc_count=8478,
    article_reported_2l_line_count=525,
    article_reported_2l_guide_count=2197,
    article_reported_2r_line_count=490,
    article_reported_2r_guide_count=1957,
    article_reported_on_target_active_arrays=168,
    article_reported_on_target_tested_arrays=169,
)


def audit_fruit_fly_cas12a_evidence(
    library_path: Path,
    genotypes_path: Path,
    source_data_path: Path,
    *,
    source: FruitFlyCas12aSource = PORT_2026_CAS12A_SOURCE,
) -> FruitFlyCas12aEvidenceAudit:
    """Verify and summarize bounded array-level in-vivo Cas12a evidence."""
    parsed, digests = _load_evidence(
        library_path,
        genotypes_path,
        source_data_path,
        source,
    )
    guide_counts = [len(array.guides) for array in parsed.arrays.values()]
    return FruitFlyCas12aEvidenceAudit(
        species_profile="fruit_fly",
        nuclease="LbCas12a-D156R",
        reagent_level="three_or_four_guide_array",
        evidence_task="in_vivo_array_level_loh_activity_observation",
        benchmark_status="usable_bounded_benchmark",
        source_id=source.source_id,
        source_reference=source.article_reference,
        source_license_note=source.license_note,
        library_url=source.library_url,
        genotypes_url=source.genotypes_url,
        source_data_url=source.source_data_url,
        library_sha256=digests[0],
        genotypes_sha256=digests[1],
        source_data_sha256=digests[2],
        source_verification="all_pinned_sources_verified",
        library_array_count=len(parsed.arrays),
        library_guide_count=sum(guide_counts),
        three_guide_array_count=guide_counts.count(3),
        four_guide_array_count=guide_counts.count(4),
        fig5i_row_count=parsed.fig5i_row_count,
        fig5i_numeric_array_observation_count=(
            parsed.fig5i_numeric_array_observation_count
        ),
        fig5i_missing_array_observation_count=(
            parsed.fig5i_missing_array_observation_count
        ),
        fig5i_control_observation_count=(
            parsed.fig5i_control_observation_count
        ),
        fig5i_unique_array_count=parsed.fig5i_unique_array_count,
        fig5j_row_count=parsed.fig5j_row_count,
        fig5j_scored_array_count=parsed.fig5j_scored_array_count,
        fig5j_disc_count=parsed.fig5j_disc_count,
        article_reported_2l_line_count=source.article_reported_2l_line_count,
        article_reported_2l_guide_count=source.article_reported_2l_guide_count,
        article_reported_2r_line_count=source.article_reported_2r_line_count,
        article_reported_2r_guide_count=source.article_reported_2r_guide_count,
        article_reported_on_target_active_arrays=(
            source.article_reported_on_target_active_arrays
        ),
        article_reported_on_target_tested_arrays=(
            source.article_reported_on_target_tested_arrays
        ),
        article_aggregate_reconstruction_status=(
            "publisher_claim_retained_not_reconstructed_without_interval_manifest"
        ),
        interval_manifest_status=(
            "unavailable_array_interval_relationship_not_reconstructed"
        ),
        predictive_adapter_available=False,
        discrimination_metrics_status=(
            "not_applicable_extreme_class_imbalance"
        ),
        warnings=_warnings(),
    )


def lookup_fruit_fly_cas12a_array(
    library_path: Path,
    genotypes_path: Path,
    source_data_path: Path,
    line_id: str,
    *,
    source: FruitFlyCas12aSource = PORT_2026_CAS12A_SOURCE,
) -> FruitFlyCas12aArrayEvidence:
    """Return source-bound evidence for one indivisible Cas12a array."""
    parsed, _ = _load_evidence(
        library_path,
        genotypes_path,
        source_data_path,
        source,
    )
    normalized = _normalize_line_id(line_id)
    if normalized not in parsed.arrays:
        raise ValueError(f"unknown HD12aCFD array {line_id!r}.")
    array = parsed.arrays[normalized]
    fig5i_scores = parsed.fig5i_scores.get(normalized, ())
    fig5i_missing = parsed.fig5i_missing.get(normalized, 0)
    fig5j_counts = parsed.fig5j_counts.get(normalized)
    fig5j_disc_count = sum(fig5j_counts) if fig5j_counts else 0
    memberships = []
    if normalized in parsed.fig5i_scores or normalized in parsed.fig5i_missing:
        memberships.append("fig5i_2l")
    if normalized in parsed.fig5j_counts:
        memberships.append("fig5j_2r")
    return FruitFlyCas12aArrayEvidence(
        species_profile="fruit_fly",
        nuclease="LbCas12a-D156R",
        source_id=source.source_id,
        source_line_id=array.source_line_id,
        primary_gene_id=array.primary_gene_id,
        primary_gene_symbol=array.primary_gene_symbol,
        component_guide_count=len(array.guides),
        component_sequence_sha256=tuple(
            sha256(guide.encode("ascii")).hexdigest()
            for guide in array.guides
        ),
        fig5i_numeric_observation_count=len(fig5i_scores),
        fig5i_missing_observation_count=fig5i_missing,
        fig5i_score_counts_0_to_3=tuple(
            fig5i_scores.count(score) for score in range(4)
        ),
        fig5i_mean_score=(
            sum(fig5i_scores) / len(fig5i_scores)
            if fig5i_scores
            else None
        ),
        fig5j_disc_count=fig5j_disc_count,
        fig5j_score_counts_0_to_3=fig5j_counts,
        fig5j_mean_score=(
            sum(score * count for score, count in enumerate(fig5j_counts))
            / fig5j_disc_count
            if fig5j_counts and fig5j_disc_count
            else None
        ),
        screen_membership="+".join(memberships) if memberships else "not_screened",
        interval_relationship="unresolved_without_interval_manifest",
        interpretation=(
            "array_level_loh_observation_with_unresolved_interval_relationship"
        ),
        calibration_status="not_a_per_guide_or_probability_calibration",
        warnings=_warnings(),
    )


def _load_evidence(
    library_path: Path,
    genotypes_path: Path,
    source_data_path: Path,
    source: FruitFlyCas12aSource,
) -> tuple[_ParsedEvidence, tuple[str, str, str]]:
    library_digest, library_bytes = _verified_bytes(
        library_path,
        source.library_sha256,
        "library",
    )
    genotypes_digest, genotypes_bytes = _verified_bytes(
        genotypes_path,
        source.genotypes_sha256,
        "genotypes",
    )
    source_data_digest, source_data_bytes = _verified_bytes(
        source_data_path,
        source.source_data_sha256,
        "source data",
    )
    arrays = _read_library(library_bytes, source)
    _verify_genotypes(genotypes_bytes)
    parsed = _read_source_data(source_data_bytes, arrays, source)
    return (
        _ParsedEvidence(arrays=arrays, **parsed),
        (library_digest, genotypes_digest, source_data_digest),
    )


def _verified_bytes(
    path: Path,
    expected_digest: str,
    label: str,
) -> tuple[str, bytes]:
    if not _SHA256_PATTERN.fullmatch(expected_digest):
        raise ValueError(f"{label} expected SHA-256 is invalid.")
    with path.open("rb") as handle:
        content = handle.read(_MAX_SOURCE_BYTES + 1)
    if not 1 <= len(content) <= _MAX_SOURCE_BYTES:
        raise ValueError(f"{label} must be between 1 byte and 20 MB.")
    digest = sha256(content).hexdigest()
    if digest != expected_digest:
        raise ValueError(f"{label} SHA-256 does not match the pinned source.")
    return digest, content


def _read_library(
    content: bytes,
    source: FruitFlyCas12aSource,
) -> dict[str, _LibraryArray]:
    try:
        stream = StringIO(content.decode("utf-8-sig"))
    except UnicodeDecodeError as error:
        raise ValueError("library must be UTF-8 CSV.") from error
    title = next(stream, "").rstrip("\r\n")
    if not title.startswith("Supplementary Table 1"):
        raise ValueError("library title row is not the qualified source.")
    reader = csv.DictReader(stream)
    if tuple(reader.fieldnames or ()) != _LIBRARY_HEADERS:
        raise ValueError("library columns do not match the qualified source.")
    arrays: dict[str, _LibraryArray] = {}
    for source_index, row in enumerate(
        islice(reader, _MAX_SOURCE_ROWS + 1),
        start=1,
    ):
        if source_index > _MAX_SOURCE_ROWS:
            raise ValueError(
                f"library exceeds the {_MAX_SOURCE_ROWS}-row safety limit."
            )
        row_number = source_index + 2
        source_line_id = str(row["Line_ID"]).strip()
        normalized = _normalize_line_id(source_line_id)
        if normalized in arrays:
            raise ValueError(f"duplicate library array at row {row_number}.")
        guides = tuple(
            str(row[f"sgRNA_{index}"]).strip().upper()
            for index in range(1, 5)
            if str(row[f"sgRNA_{index}"]).strip()
        )
        if len(guides) not in {3, 4} or any(
            not _GUIDE_PATTERN.fullmatch(guide) for guide in guides
        ):
            raise ValueError(
                f"library row {row_number} must contain three or four "
                "valid 23-nt spacers."
            )
        gene_id = str(row["primary_gene"]).strip()
        if not _GENE_ID_PATTERN.fullmatch(gene_id):
            raise ValueError(f"library row {row_number} has an invalid FlyBase ID.")
        gene_symbol = str(row["primary_gene_symbol"]).strip() or None
        arrays[normalized] = _LibraryArray(
            source_line_id=source_line_id,
            primary_gene_id=gene_id,
            primary_gene_symbol=gene_symbol,
            guides=guides,
        )
    if len(arrays) != source.expected_library_array_count:
        raise ValueError("library array count does not match the source profile.")
    if sum(len(array.guides) for array in arrays.values()) != (
        source.expected_library_guide_count
    ):
        raise ValueError("library guide count does not match the source profile.")
    return arrays


def _verify_genotypes(content: bytes) -> None:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        if "Sheet1" not in workbook.sheetnames:
            raise ValueError("genotypes workbook is missing Sheet1.")
        rows = workbook["Sheet1"].iter_rows(values_only=True)
        title = next(rows, ())
        if not title or not str(title[0]).startswith("Supplementary Table 3"):
            raise ValueError("genotypes title row is not the qualified source.")
        headers = tuple(next(rows, ())[:5])
        expected = (
            "Stock_Name",
            "Genotype",
            "Used in",
            "Internal_Transgene_IDs_Source",
            "Comment",
        )
        if headers != expected:
            raise ValueError(
                "genotypes columns do not match the qualified source."
            )
        matches = [
            row
            for row in islice(rows, _MAX_SOURCE_ROWS)
            if row[0] == "HD12aCFD sgRNA lines"
        ]
        if len(matches) != 1 or "attP40" not in str(matches[0][1]):
            raise ValueError(
                "genotypes workbook does not uniquely define the HD12aCFD "
                "attP40 stock."
            )
    finally:
        workbook.close()


def _read_source_data(
    content: bytes,
    arrays: dict[str, _LibraryArray],
    source: FruitFlyCas12aSource,
) -> dict[str, object]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        for sheet_name in ("Fig5i", "Fig5j"):
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"source data workbook is missing {sheet_name}.")
        fig5i = _read_fig5i(workbook["Fig5i"], arrays)
        fig5j = _read_fig5j(workbook["Fig5j"], arrays)
    finally:
        workbook.close()
    checks = (
        (fig5i["row_count"], source.expected_fig5i_row_count, "Fig5i row"),
        (
            fig5i["numeric_count"],
            source.expected_fig5i_numeric_array_observation_count,
            "Fig5i numeric observation",
        ),
        (
            fig5i["missing_count"],
            source.expected_fig5i_missing_array_observation_count,
            "Fig5i missing observation",
        ),
        (
            fig5i["control_count"],
            source.expected_fig5i_control_observation_count,
            "Fig5i control observation",
        ),
        (
            len(set(fig5i["scores"]) | set(fig5i["missing"])),
            source.expected_fig5i_unique_array_count,
            "Fig5i unique array",
        ),
        (fig5j["row_count"], source.expected_fig5j_row_count, "Fig5j row"),
        (
            fig5j["scored_count"],
            source.expected_fig5j_scored_array_count,
            "Fig5j scored array",
        ),
        (
            fig5j["disc_count"],
            source.expected_fig5j_disc_count,
            "Fig5j disc",
        ),
    )
    for actual, expected, label in checks:
        if actual != expected:
            raise ValueError(f"{label} count does not match the source profile.")
    return {
        "fig5i_scores": {
            line_id: tuple(scores)
            for line_id, scores in fig5i["scores"].items()
        },
        "fig5i_missing": dict(fig5i["missing"]),
        "fig5j_counts": dict(fig5j["counts"]),
        "fig5i_row_count": fig5i["row_count"],
        "fig5i_numeric_array_observation_count": fig5i["numeric_count"],
        "fig5i_missing_array_observation_count": fig5i["missing_count"],
        "fig5i_control_observation_count": fig5i["control_count"],
        "fig5i_unique_array_count": len(
            set(fig5i["scores"]) | set(fig5i["missing"])
        ),
        "fig5j_row_count": fig5j["row_count"],
        "fig5j_scored_array_count": fig5j["scored_count"],
        "fig5j_disc_count": fig5j["disc_count"],
    }


def _read_fig5i(sheet, arrays: dict[str, _LibraryArray]) -> dict[str, object]:
    rows = sheet.iter_rows(values_only=True)
    if tuple(next(rows, ())) != _FIG5I_HEADERS:
        raise ValueError("Fig5i columns do not match the qualified source.")
    scores: dict[str, list[int]] = {}
    missing: dict[str, int] = {}
    row_count = numeric_count = missing_count = control_count = 0
    for row in islice(rows, _MAX_SOURCE_ROWS):
        row_count += 1
        line_id = str(row[0]).strip()
        if line_id.upper() == "EMPTYCFD8":
            control_count += 1
            continue
        normalized = _normalize_line_id(line_id)
        if normalized not in arrays:
            raise ValueError(f"Fig5i contains unknown array {line_id!r}.")
        score = row[1]
        if score == "NA":
            missing[normalized] = missing.get(normalized, 0) + 1
            missing_count += 1
        elif isinstance(score, int) and not isinstance(score, bool) and 0 <= score <= 3:
            scores.setdefault(normalized, []).append(score)
            numeric_count += 1
        else:
            raise ValueError(f"Fig5i has an invalid score for {line_id}.")
    return {
        "scores": scores,
        "missing": missing,
        "row_count": row_count,
        "numeric_count": numeric_count,
        "missing_count": missing_count,
        "control_count": control_count,
    }


def _read_fig5j(sheet, arrays: dict[str, _LibraryArray]) -> dict[str, object]:
    rows = sheet.iter_rows(values_only=True)
    if tuple(next(rows, ())) != _FIG5J_HEADERS:
        raise ValueError("Fig5j columns do not match the qualified source.")
    counts: dict[str, tuple[int, int, int, int] | None] = {}
    row_count = scored_count = disc_count = 0
    for row in islice(rows, _MAX_SOURCE_ROWS):
        row_count += 1
        line_id = str(row[0]).strip()
        normalized = _normalize_line_id(line_id)
        if normalized not in arrays:
            raise ValueError(f"Fig5j contains unknown array {line_id!r}.")
        if normalized in counts:
            raise ValueError(f"Fig5j contains duplicate array {line_id!r}.")
        raw_counts = row[1:5]
        if all(value is None for value in raw_counts):
            counts[normalized] = None
            continue
        normalized_counts = tuple(
            0 if value is None else value for value in raw_counts
        )
        if any(
            isinstance(value, bool)
            or not isinstance(value, int)
            or value < 0
            for value in normalized_counts
        ):
            raise ValueError(f"Fig5j has invalid category counts for {line_id}.")
        total = sum(normalized_counts)
        if total < 1:
            raise ValueError(f"Fig5j has an empty scored row for {line_id}.")
        counts[normalized] = normalized_counts
        scored_count += 1
        disc_count += total
    return {
        "counts": counts,
        "row_count": row_count,
        "scored_count": scored_count,
        "disc_count": disc_count,
    }


def _normalize_line_id(value: str) -> str:
    candidate = str(value).strip().upper()
    match = _LINE_ID_PATTERN.fullmatch(candidate)
    if not match:
        raise ValueError(f"invalid HD12aCFD array identifier {value!r}.")
    return f"HD12ACFD{int(match.group(1)):04d}"


def _warnings() -> tuple[str, ...]:
    return (
        "Every observation belongs to a three- or four-guide array and cannot "
        "be assigned to an individual component guide.",
        "The publication reports 168 active arrays among 169 in-window arrays; "
        "this class imbalance cannot support discrimination or calibration.",
        "The monitored LOH intervals cover part of chromosome 2, not "
        "genome-wide off-target recall.",
        "The source files audited here do not reconstruct each target's "
        "relationship to the monitored interval; a zero LOH score must not be "
        "interpreted as an inactive array.",
        "LOH activity does not predict repair spectrum, phenotype, animal "
        "welfare, or safety in another stock, tissue, stage, or nuclease.",
    )
