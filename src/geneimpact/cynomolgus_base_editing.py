"""Pinned cynomolgus-embryo base-editing transfer evidence."""

from __future__ import annotations

from contextlib import contextmanager
from dataclasses import dataclass
from hashlib import sha256
from io import BytesIO
import json
import math
from pathlib import Path
import re
from typing import Any, Iterator, Mapping, Sequence
import warnings

from openpyxl import load_workbook

from .species import PROFILES


GENEIMPACT_VERSION = "0.16.1"
CYNOMOLGUS_BASE_EDITING_REFERENCE = (
    "https://doi.org/10.1038/s41467-020-16173-0"
)
CYNOMOLGUS_BASE_EDITING_TARGET_SITES_URL = (
    "https://static-content.springer.com/esm/art%3A10.1038%2F"
    "s41467-020-16173-0/MediaObjects/"
    "41467_2020_16173_MOESM1_ESM.xlsx"
)
CYNOMOLGUS_BASE_EDITING_SOURCE_DATA_URL = (
    "https://static-content.springer.com/esm/art%3A10.1038%2F"
    "s41467-020-16173-0/MediaObjects/"
    "41467_2020_16173_MOESM11_ESM.xlsx"
)
CYNOMOLGUS_BASE_EDITING_TARGET_SITES_SHA256 = (
    "d736c4a0607c9590418883fa5f38f68f1ee5a319b7e3344ea00dcbfb1be08de3"
)
CYNOMOLGUS_BASE_EDITING_SOURCE_DATA_SHA256 = (
    "efe82aca05dffc4ba96df0ee177a90128f5e466d6a6dc3b282c9c817d05c763a"
)
CYNOMOLGUS_BASE_EDITING_DATA_SPLIT_IDENTIFIER = (
    "zhang-2020-fixed-external-transfer-v1"
)
CYNOMOLGUS_BASE_EDITING_TARGET_ASSEMBLY_ACCESSION = PROFILES[
    "cynomolgus_macaque"
].assembly_accession
CYNOMOLGUS_BASE_EDITING_TARGET_GENOME_BUILD = PROFILES[
    "cynomolgus_macaque"
].genome_build

_MAX_TARGET_SITES_BYTES = 2_000_000
_MAX_SOURCE_DATA_BYTES = 64_000_000
_SHA256_PATTERN = re.compile(r"^[0-9a-f]{64}$")
_TARGET_SEQUENCE_PATTERN = re.compile(r"^[ACGT]{20,30}$")
_COUNT_PATTERN = re.compile(r"^\s*([0-9]+)(?:\s*\(|\s*$)")
_TARGET_HEADERS = (
    "Target site",
    "Gene",
    "Type",
    "Sequence",
    "Chromosome",
    "Position",
    "Direction",
)


@dataclass(frozen=True)
class CynomolgusBaseEditingBlock:
    """One target table within a shared embryo-injection context."""

    context_id: str
    editor: str
    multiplex_guide_count: int
    target_site_id: str
    gene: str
    conversion: str
    target_bases: tuple[str, ...]
    label_row: int
    header_row: int
    first_data_row: int
    last_data_row: int
    intended_count_columns: tuple[int, ...]
    source_label: str | None = None


@dataclass(frozen=True)
class CynomolgusBaseEditingSource:
    """Checksums and structural expectations for a qualified source pair."""

    source_id: str
    article_reference: str
    target_sites_url: str
    source_data_url: str
    target_sites_sha256: str
    source_data_sha256: str
    source_genome_build: str
    expected_candidate_site_count: int
    expected_on_target_site_count: int
    expected_record_count: int
    expected_context_count: int
    expected_embryo_base_observation_count: int
    expected_clone_denominator_count: int
    blocks: tuple[CynomolgusBaseEditingBlock, ...]
    license_note: str = (
        "Article and supplementary files are licensed CC BY 4.0; "
        "GeneImpact AI does not bundle the publisher workbooks."
    )


@dataclass(frozen=True)
class _ObservedBaseRecord:
    record_id: str
    context_id: str
    editor: str
    multiplex_guide_count: int
    target_site_id: str
    gene: str
    conversion: str
    target_base: str
    target_sequence: str
    intended_count: int
    clone_denominator: int
    embryo_count: int


@dataclass(frozen=True)
class CynomolgusBaseEditingTransferMetrics:
    """Metrics that do not compare arbitrary scores across contexts."""

    within_context_candidate_pair_count: int
    within_context_eligible_pair_count: int
    within_context_observation_tie_pair_count: int
    within_context_prediction_tie_pair_count: int
    within_context_pair_count: int
    within_context_concordant_pair_count: int
    within_context_weighted_concordant_score: float
    within_context_pairwise_accuracy: float | None
    within_context_prediction_coverage: float | None
    mean_absolute_error: float | None
    root_mean_squared_error: float | None


@dataclass(frozen=True)
class CynomolgusBaseEditingTransferReport:
    predictor: str
    predictor_version: str
    species_profile: str
    evaluation_status: str
    use: str
    predictive_adapter_available: bool
    source_id: str
    source_reference: str
    source_target_sites_url: str
    source_data_url: str
    source_license_note: str
    source_genome_build: str
    current_registered_genome_build: str
    source_assembly_accession: str
    target_assembly_accession: str
    liftover_status: str
    publisher_target_sequence_record_verified: bool
    target_sequence_verified_on_source_assembly: bool
    target_sequence_verified_on_target: bool
    target_sites_sha256: str
    source_data_sha256: str
    source_verification: str
    evidence_snapshot_identifier: str
    submitted_code_revision: str
    code_revision_verified: bool
    geneimpact_version: str
    evaluator_code_revision: str
    evaluator_code_revision_verified: bool
    evaluator_code_revision_status: str
    data_split_identifier: str
    study_context: str
    population_or_strain: str
    functional_annotation_status: str
    exclusion_rules: tuple[str, ...]
    target_site_count: int
    record_count: int
    context_count: int
    comparison_stratum_count: int
    embryo_base_observation_count: int
    clone_denominator_count: int
    editors: tuple[str, ...]
    score_semantics: str
    prediction_target: str
    sequence_basis: str
    training_overlap_status: str
    training_overlap_evidence_reference: str
    confidence_interval_status: str
    confidence_interval_note: str
    prediction_submission_sha256: str
    independence_verified: bool
    independence_interpretation: str
    metrics: CynomolgusBaseEditingTransferMetrics
    warnings: tuple[str, ...]


_ZHANG_2020_BLOCKS = (
    CynomolgusBaseEditingBlock(
        "single_be3_fah_e4",
        "BE3",
        1,
        "FAH-E4",
        "FAH",
        "C_to_T",
        ("C5", "C6"),
        3,
        4,
        5,
        20,
        (3, 5),
        source_label="FAH",
    ),
    CynomolgusBaseEditingBlock(
        "single_abe710_app",
        "ABE7.10",
        1,
        "APP",
        "APP",
        "A_to_G",
        ("A5", "A7"),
        22,
        23,
        24,
        32,
        (3, 5),
    ),
    CynomolgusBaseEditingBlock(
        "double_be3_fah_e7_e9sg1",
        "BE3",
        2,
        "FAH-E7",
        "FAH",
        "C_to_T",
        ("C4", "C7"),
        36,
        37,
        38,
        46,
        (3, 5),
    ),
    CynomolgusBaseEditingBlock(
        "double_be3_fah_e7_e9sg1",
        "BE3",
        2,
        "FAH-E9-sg1",
        "FAH",
        "C_to_T",
        ("C4", "C6"),
        48,
        49,
        50,
        58,
        (3, 5),
    ),
    CynomolgusBaseEditingBlock(
        "triple_be3_fah_e4_e9sg2_e14",
        "BE3",
        3,
        "FAH-E4",
        "FAH",
        "C_to_T",
        ("C5", "C6"),
        61,
        62,
        63,
        72,
        (3, 5),
    ),
    CynomolgusBaseEditingBlock(
        "triple_be3_fah_e4_e9sg2_e14",
        "BE3",
        3,
        "FAH-E9-sg2",
        "FAH",
        "C_to_T",
        ("C3", "C4", "C6", "C8", "C9"),
        74,
        75,
        76,
        85,
        (3, 5, 7, 9, 11),
    ),
    CynomolgusBaseEditingBlock(
        "triple_be3_fah_e4_e9sg2_e14",
        "BE3",
        3,
        "FAH-E14",
        "FAH",
        "C_to_T",
        ("C3", "C6", "C7"),
        87,
        88,
        89,
        98,
        (3, 5, 7),
    ),
    CynomolgusBaseEditingBlock(
        "double_abe710_hbb_tp53",
        "ABE7.10",
        2,
        "HBB",
        "HBB",
        "A_to_G",
        ("A4", "A5"),
        101,
        102,
        103,
        107,
        (3, 5),
    ),
    CynomolgusBaseEditingBlock(
        "double_abe710_hbb_tp53",
        "ABE7.10",
        2,
        "TP53",
        "TP53",
        "A_to_G",
        ("A7",),
        109,
        110,
        111,
        115,
        (3,),
    ),
    CynomolgusBaseEditingBlock(
        "triple_sakkh_be3_abe710_emx1_fancf_brca1",
        "SaKKH-BE3",
        3,
        "EMX1",
        "EMX1",
        "C_to_T",
        ("C7", "C8"),
        119,
        120,
        121,
        128,
        (3, 5),
    ),
    CynomolgusBaseEditingBlock(
        "triple_sakkh_be3_abe710_emx1_fancf_brca1",
        "SaKKH-BE3",
        3,
        "FANCF",
        "FANCF",
        "C_to_T",
        ("C6", "C7", "C8", "C12"),
        130,
        131,
        132,
        139,
        (3, 5, 7, 9),
    ),
    CynomolgusBaseEditingBlock(
        "triple_sakkh_be3_abe710_emx1_fancf_brca1",
        "ABE7.10",
        3,
        "BRCA1",
        "BRCA1",
        "A_to_G",
        ("A2", "A3", "A5"),
        141,
        142,
        143,
        150,
        (3, 5, 7),
    ),
)


ZHANG_2020_CYNOMOLGUS_BASE_EDITING_SOURCE = CynomolgusBaseEditingSource(
    source_id="zhang-2020-cynomolgus-embryo-base-editing",
    article_reference=CYNOMOLGUS_BASE_EDITING_REFERENCE,
    target_sites_url=CYNOMOLGUS_BASE_EDITING_TARGET_SITES_URL,
    source_data_url=CYNOMOLGUS_BASE_EDITING_SOURCE_DATA_URL,
    target_sites_sha256=CYNOMOLGUS_BASE_EDITING_TARGET_SITES_SHA256,
    source_data_sha256=CYNOMOLGUS_BASE_EDITING_SOURCE_DATA_SHA256,
    source_genome_build="GCF_000364345.1",
    expected_candidate_site_count=66,
    expected_on_target_site_count=11,
    expected_record_count=30,
    expected_context_count=6,
    expected_embryo_base_observation_count=273,
    expected_clone_denominator_count=8296,
    blocks=_ZHANG_2020_BLOCKS,
)


def prepare_cynomolgus_base_editing_transfer_template(
    target_sites_path: Path,
    source_data_path: Path,
    *,
    source: CynomolgusBaseEditingSource = (
        ZHANG_2020_CYNOMOLGUS_BASE_EDITING_SOURCE
    ),
) -> dict[str, object]:
    """Build a sequence-redacted external-prediction submission template."""
    records, digests = _load_records(target_sites_path, source_data_path, source)
    return {
        "schema_version": (
            "geneimpact.cynomolgus_base_editing_transfer_predictions.v1"
        ),
        "source": {
            "source_id": source.source_id,
            "reference": source.article_reference,
            "target_sites_sha256": digests[0],
            "source_data_sha256": digests[1],
            "source_genome_build": source.source_genome_build,
            "source_assembly_accession": source.source_genome_build,
            "target_assembly_accession": (
                CYNOMOLGUS_BASE_EDITING_TARGET_ASSEMBLY_ACCESSION
            ),
            "liftover_status": "not_performed",
            "publisher_target_sequence_record_verified": True,
            "target_sequence_verified_on_source_assembly": False,
            "target_sequence_verified_on_target": False,
            "data_split_identifier": (
                CYNOMOLGUS_BASE_EDITING_DATA_SPLIT_IDENTIFIER
            ),
            "license_note": source.license_note,
        },
        "prediction": {
            "name": "REPLACE_WITH_PREDICTOR_NAME",
            "version": "REPLACE_WITH_VERSION_OR_COMMIT",
            "submitted_code_revision": "REPLACE_WITH_CODE_COMMIT",
            "score_direction": "higher_is_more_edited",
            "score_semantics": "ranking_score",
            "prediction_target": (
                "pooled_intended_base_conversion_fraction"
            ),
            "sequence_basis": "publisher_target_sequence_with_pam",
            "training_overlap_status": "unknown",
            "evidence_reference": "REPLACE_WITH_MODEL_OR_RUN_REFERENCE",
        },
        "records": [
            {
                "record_id": record.record_id,
                "context_id": record.context_id,
                "editor": record.editor,
                "multiplex_guide_count": record.multiplex_guide_count,
                "target_site_id": record.target_site_id,
                "gene": record.gene,
                "conversion": record.conversion,
                "target_base": record.target_base,
                "target_sequence_sha256": sha256(
                    record.target_sequence.encode("ascii")
                ).hexdigest(),
                "target_sequence_length": len(record.target_sequence),
                "predicted_score": None,
            }
            for record in records
        ],
        "instructions": (
            "Fill immutable predictor metadata and every predicted_score. "
            "Scores are evaluated only as a retrospective cynomolgus-embryo "
            "transfer benchmark; declare training overlap and do not alter "
            "record identities, contexts, editors, target bases, or hashes."
        ),
    }


def evaluate_cynomolgus_base_editing_transfer(
    target_sites_path: Path,
    source_data_path: Path,
    predictions: Mapping[str, Any],
    *,
    source: CynomolgusBaseEditingSource = (
        ZHANG_2020_CYNOMOLGUS_BASE_EDITING_SOURCE
    ),
) -> CynomolgusBaseEditingTransferReport:
    """Evaluate external scores without promoting a macaque predictor."""
    records, digests = _load_records(target_sites_path, source_data_path, source)
    metadata, scores = _validate_predictions(predictions, records, source)
    observed = {
        record.record_id: record.intended_count / record.clone_denominator
        for record in records
    }
    pairwise = _within_context_concordance(
        records,
        scores,
        observed,
    )
    expected_fractions = (
        metadata["score_semantics"] == "expected_edit_fraction"
    )
    absolute_errors: list[float] = []
    squared_errors: list[float] = []
    if expected_fractions:
        absolute_errors = [
            abs(scores[record.record_id] - observed[record.record_id])
            for record in records
        ]
        squared_errors = [
            (scores[record.record_id] - observed[record.record_id]) ** 2
            for record in records
        ]
    overlap_is_verified = metadata["training_overlap_status"] == (
        "declared_no_overlap"
    )
    evaluation_status = (
        "retrospective_external_transfer_benchmark"
        if overlap_is_verified
        else "descriptive_evaluation_with_unverified_overlap"
    )
    use = (
        "bounded_external_transfer_benchmark_only"
        if overlap_is_verified
        else "descriptive_only_unverified_overlap"
    )
    return CynomolgusBaseEditingTransferReport(
        predictor=metadata["name"],
        predictor_version=metadata["version"],
        species_profile="cynomolgus_macaque",
        evaluation_status=evaluation_status,
        use=use,
        predictive_adapter_available=False,
        source_id=source.source_id,
        source_reference=source.article_reference,
        source_target_sites_url=source.target_sites_url,
        source_data_url=source.source_data_url,
        source_license_note=source.license_note,
        source_genome_build=source.source_genome_build,
        current_registered_genome_build=CYNOMOLGUS_BASE_EDITING_TARGET_GENOME_BUILD,
        source_assembly_accession=source.source_genome_build,
        target_assembly_accession=CYNOMOLGUS_BASE_EDITING_TARGET_ASSEMBLY_ACCESSION,
        liftover_status="not_performed",
        publisher_target_sequence_record_verified=True,
        target_sequence_verified_on_source_assembly=False,
        target_sequence_verified_on_target=False,
        target_sites_sha256=digests[0],
        source_data_sha256=digests[1],
        source_verification="pinned_workbooks_verified",
        evidence_snapshot_identifier=source.source_id,
        submitted_code_revision=metadata["submitted_code_revision"],
        code_revision_verified=False,
        geneimpact_version=GENEIMPACT_VERSION,
        evaluator_code_revision=(
            "sha256:" + sha256(Path(__file__).read_bytes()).hexdigest()
        ),
        evaluator_code_revision_verified=True,
        evaluator_code_revision_status="module_source_sha256",
        data_split_identifier=CYNOMOLGUS_BASE_EDITING_DATA_SPLIT_IDENTIFIER,
        study_context=(
            "cynomolgus zygote editor mRNA/T7 sgRNA cytoplasmic microinjection "
            "10–12 hours after ICSI with pooled Sanger-clone genotyping"
        ),
        population_or_strain=(
            "source colony; geographic population not resolved in the source"
        ),
        functional_annotation_status="not_applicable_sequence_only",
        exclusion_rules=(
            "known training overlap is rejected",
            "observation ties are excluded from strict ranking pairs",
            "scores are never compared across editor/context strata",
            "source labels and raw target sequences are excluded from reports",
        ),
        target_site_count=len({record.target_site_id for record in records}),
        record_count=len(records),
        context_count=len({record.context_id for record in records}),
        comparison_stratum_count=len(
            {(record.context_id, record.editor) for record in records}
        ),
        embryo_base_observation_count=sum(
            record.embryo_count for record in records
        ),
        clone_denominator_count=sum(
            record.clone_denominator for record in records
        ),
        editors=tuple(sorted({record.editor for record in records})),
        score_semantics=metadata["score_semantics"],
        prediction_target=metadata["prediction_target"],
        sequence_basis=metadata["sequence_basis"],
        training_overlap_status=metadata["training_overlap_status"],
        training_overlap_evidence_reference=metadata["evidence_reference"],
        confidence_interval_status=(
            "unavailable_correlated_embryo_clustered_source"
        ),
        confidence_interval_note=(
            "No interval is reported: embryo, target-base, and clone outcomes "
            "are clustered within injection contexts and the benchmark has "
            "too few independent units for a defensible interval."
        ),
        prediction_submission_sha256=sha256(
            json.dumps(
                predictions,
                sort_keys=True,
                separators=(",", ":"),
                ensure_ascii=False,
            ).encode("utf-8")
        ).hexdigest(),
        independence_verified=False,
        independence_interpretation=(
            "Training-overlap status is submitter-declared and has not been "
            "reproduced by GeneImpact AI."
        ),
        metrics=CynomolgusBaseEditingTransferMetrics(
            within_context_candidate_pair_count=pairwise.candidate_pair_count,
            within_context_eligible_pair_count=pairwise.eligible_pair_count,
            within_context_observation_tie_pair_count=(
                pairwise.observation_tie_pair_count
            ),
            within_context_prediction_tie_pair_count=(
                pairwise.prediction_tie_pair_count
            ),
            within_context_pair_count=pairwise.strict_pair_count,
            within_context_concordant_pair_count=pairwise.concordant_pair_count,
            within_context_weighted_concordant_score=(
                pairwise.weighted_concordant_score
            ),
            within_context_pairwise_accuracy=pairwise.accuracy,
            within_context_prediction_coverage=pairwise.prediction_coverage,
            mean_absolute_error=(
                sum(absolute_errors) / len(records)
                if expected_fractions
                else None
            ),
            root_mean_squared_error=(
                math.sqrt(sum(squared_errors) / len(records))
                if expected_fractions
                else None
            ),
        ),
        warnings=(
            "Only 11 target sites and 30 target-base/context records are "
            "represented; this dataset is too small for species-level "
            "training or calibration.",
            "Embryos, target bases, and clone counts within one injection "
            "context are not independent observations.",
            "Ranking agreement is computed only within a shared editor and "
            "multiplex injection context; arbitrary ranking scores are not "
            "compared across contexts.",
            "The source used legacy assembly GCF_000364345.1, not the current "
            "registered T2T-MFA8v1.1 assembly.",
            "Training-overlap status is self-declared; this report is not an "
            "independently verified held-out test. Unknown overlap is "
            "descriptive only and is not labelled external validation.",
            "The source uses zygote mRNA/sgRNA microinjection and pooled "
            "Sanger-clone fractions; results do not transfer automatically "
            "to another editor, delivery, developmental stage, population, "
            "laboratory, or assay.",
            "This benchmark does not establish phenotype, repair-spectrum, "
            "off-target safety, animal welfare, or prospective performance.",
        ),
    )


def _validate_predictions(
    document: Mapping[str, Any],
    records: tuple[_ObservedBaseRecord, ...],
    source: CynomolgusBaseEditingSource,
) -> tuple[dict[str, str], dict[str, float]]:
    if document.get("schema_version") != (
        "geneimpact.cynomolgus_base_editing_transfer_predictions.v1"
    ):
        raise ValueError("prediction schema_version is not supported.")
    source_metadata = document.get("source")
    metadata = document.get("prediction")
    raw_records = document.get("records")
    if not isinstance(source_metadata, Mapping):
        raise ValueError("source must be an object.")
    expected_source = {
        "source_id": source.source_id,
        "reference": source.article_reference,
        "target_sites_sha256": source.target_sites_sha256,
        "source_data_sha256": source.source_data_sha256,
        "source_genome_build": source.source_genome_build,
        "source_assembly_accession": source.source_genome_build,
        "target_assembly_accession": CYNOMOLGUS_BASE_EDITING_TARGET_ASSEMBLY_ACCESSION,
        "liftover_status": "not_performed",
        "publisher_target_sequence_record_verified": True,
        "target_sequence_verified_on_source_assembly": False,
        "target_sequence_verified_on_target": False,
        "data_split_identifier": CYNOMOLGUS_BASE_EDITING_DATA_SPLIT_IDENTIFIER,
        "license_note": source.license_note,
    }
    if any(
        source_metadata.get(key) != value
        for key, value in expected_source.items()
    ):
        raise ValueError("prediction source metadata does not match the source.")
    if not isinstance(metadata, Mapping):
        raise ValueError("prediction must be an object.")
    if not isinstance(raw_records, Sequence) or isinstance(
        raw_records,
        (str, bytes),
    ):
        raise ValueError("records must be a list.")
    values = {
        key: str(metadata.get(key, "")).strip()
        for key in (
            "name",
            "version",
            "submitted_code_revision",
            "score_direction",
            "score_semantics",
            "prediction_target",
            "sequence_basis",
            "training_overlap_status",
            "evidence_reference",
        )
    }
    if (
        not values["name"]
        or values["name"].startswith("REPLACE_WITH")
        or not values["version"]
        or values["version"].startswith("REPLACE_WITH")
        or not values["submitted_code_revision"]
        or values["submitted_code_revision"].startswith("REPLACE_WITH")
        or not values["evidence_reference"]
        or values["evidence_reference"].startswith("REPLACE_WITH")
    ):
        raise ValueError(
            "predictor name, version, submitted code revision, and evidence "
            "are required."
        )
    if values["score_direction"] != "higher_is_more_edited":
        raise ValueError("score_direction must be higher_is_more_edited.")
    if values["score_semantics"] not in {
        "ranking_score",
        "expected_edit_fraction",
    }:
        raise ValueError(
            "score_semantics must be ranking_score or expected_edit_fraction."
        )
    if values["prediction_target"] != (
        "pooled_intended_base_conversion_fraction"
    ):
        raise ValueError(
            "prediction_target must be "
            "pooled_intended_base_conversion_fraction."
        )
    if values["sequence_basis"] != "publisher_target_sequence_with_pam":
        raise ValueError(
            "sequence_basis must be publisher_target_sequence_with_pam."
        )
    if values["training_overlap_status"] not in {
        "unknown",
        "declared_no_overlap",
    }:
        raise ValueError(
            "known training overlap is not eligible for transfer evaluation."
        )
    expected = {record.record_id: record for record in records}
    if len(raw_records) != len(expected):
        raise ValueError("prediction record count does not match the source.")
    scores: dict[str, float] = {}
    for index, raw_record in enumerate(raw_records, start=1):
        if not isinstance(raw_record, Mapping):
            raise ValueError(f"prediction record {index} must be an object.")
        record_id = str(raw_record.get("record_id", "")).strip()
        if record_id not in expected or record_id in scores:
            raise ValueError(
                f"prediction record {index} has an unknown or duplicate ID."
            )
        record = expected[record_id]
        immutable = {
            "context_id": record.context_id,
            "editor": record.editor,
            "multiplex_guide_count": record.multiplex_guide_count,
            "target_site_id": record.target_site_id,
            "gene": record.gene,
            "conversion": record.conversion,
            "target_base": record.target_base,
            "target_sequence_sha256": sha256(
                record.target_sequence.encode("ascii")
            ).hexdigest(),
            "target_sequence_length": len(record.target_sequence),
        }
        if any(raw_record.get(key) != value for key, value in immutable.items()):
            raise ValueError(
                f"prediction record {record_id} identity was altered."
            )
        score = _finite_number(
            raw_record.get("predicted_score"),
            f"predicted_score for {record_id}",
        )
        if (
            values["score_semantics"] == "expected_edit_fraction"
            and not 0 <= score <= 1
        ):
            raise ValueError(
                f"predicted_score for {record_id} must be between 0 and 1."
            )
        scores[record_id] = score
    return values, scores


@dataclass(frozen=True)
class _PairwiseConcordance:
    candidate_pair_count: int
    eligible_pair_count: int
    observation_tie_pair_count: int
    prediction_tie_pair_count: int
    strict_pair_count: int
    concordant_pair_count: int
    weighted_concordant_score: float
    accuracy: float | None
    prediction_coverage: float | None


def _within_context_concordance(
    records: tuple[_ObservedBaseRecord, ...],
    scores: Mapping[str, float],
    observed: Mapping[str, float],
) -> _PairwiseConcordance:
    groups: dict[tuple[str, str], list[_ObservedBaseRecord]] = {}
    for record in records:
        groups.setdefault(
            (record.context_id, record.editor),
            [],
        ).append(record)
    candidate_pairs = 0
    observation_ties = 0
    prediction_ties = 0
    strict_pairs = 0
    concordant = 0
    weighted_concordant = 0.0
    for group in groups.values():
        for left_index, left in enumerate(group):
            for right in group[left_index + 1 :]:
                candidate_pairs += 1
                prediction_delta = (
                    scores[left.record_id] - scores[right.record_id]
                )
                observation_delta = (
                    observed[left.record_id] - observed[right.record_id]
                )
                if observation_delta == 0:
                    observation_ties += 1
                    continue
                if prediction_delta == 0:
                    prediction_ties += 1
                    weighted_concordant += 0.5
                    continue
                strict_pairs += 1
                is_concordant = math.copysign(1, prediction_delta) == math.copysign(
                    1, observation_delta
                )
                concordant += int(is_concordant)
                weighted_concordant += int(is_concordant)
    eligible_pairs = candidate_pairs - observation_ties
    return _PairwiseConcordance(
        candidate_pair_count=candidate_pairs,
        eligible_pair_count=eligible_pairs,
        observation_tie_pair_count=observation_ties,
        prediction_tie_pair_count=prediction_ties,
        strict_pair_count=strict_pairs,
        concordant_pair_count=concordant,
        weighted_concordant_score=weighted_concordant,
        accuracy=(
            weighted_concordant / eligible_pairs
            if eligible_pairs
            else None
        ),
        prediction_coverage=(
            strict_pairs / eligible_pairs if eligible_pairs else None
        ),
    )


def _finite_number(value: object, label: str) -> float:
    if isinstance(value, bool):
        raise ValueError(f"{label} must be numeric.")
    try:
        number = float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{label} must be numeric.") from error
    if not math.isfinite(number):
        raise ValueError(f"{label} must be finite.")
    return number


def _load_records(
    target_sites_path: Path,
    source_data_path: Path,
    source: CynomolgusBaseEditingSource,
) -> tuple[tuple[_ObservedBaseRecord, ...], tuple[str, str]]:
    target_digest, target_bytes = _verified_bytes(
        target_sites_path,
        source.target_sites_sha256,
        _MAX_TARGET_SITES_BYTES,
        "target-sites workbook",
    )
    source_digest, source_bytes = _verified_bytes(
        source_data_path,
        source.source_data_sha256,
        _MAX_SOURCE_DATA_BYTES,
        "source-data workbook",
    )
    target_sequences = _read_target_sequences(target_bytes, source)
    records = _read_observations(source_bytes, target_sequences, source)
    return records, (target_digest, source_digest)


def _verified_bytes(
    path: Path,
    expected_digest: str,
    maximum_bytes: int,
    label: str,
) -> tuple[str, bytes]:
    if not _SHA256_PATTERN.fullmatch(expected_digest):
        raise ValueError(f"{label} expected SHA-256 is invalid.")
    with path.open("rb") as handle:
        content = handle.read(maximum_bytes + 1)
    if not 1 <= len(content) <= maximum_bytes:
        raise ValueError(
            f"{label} must be between 1 byte and {maximum_bytes} bytes."
        )
    digest = sha256(content).hexdigest()
    if digest != expected_digest:
        raise ValueError(f"{label} SHA-256 does not match the pinned source.")
    return digest, content


def _read_target_sequences(
    content: bytes,
    source: CynomolgusBaseEditingSource,
) -> dict[str, str]:
    with _publisher_workbook(content) as workbook:
        if "sheet1" not in workbook.sheetnames:
            raise ValueError("target-sites workbook is missing sheet1.")
        sheet = workbook["sheet1"]
        if not str(sheet.cell(1, 1).value).startswith(
            "Supplementary Data 1 Off-target sites"
        ):
            raise ValueError("target-sites workbook title is not qualified.")
        headers = tuple(
            sheet.cell(2, column).value
            for column in range(1, len(_TARGET_HEADERS) + 1)
        )
        if headers != _TARGET_HEADERS:
            raise ValueError("target-sites workbook columns do not match.")
        sequences: dict[str, str] = {}
        candidate_count = 0
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row[0]:
                continue
            candidate_count += 1
            target_id = str(row[0]).strip()
            if not target_id.endswith("-On"):
                continue
            normalized_id = target_id[:-3]
            sequence = re.sub(r"\s+", "", str(row[3]).upper())
            if not _TARGET_SEQUENCE_PATTERN.fullmatch(sequence):
                raise ValueError(
                    f"invalid publisher target sequence for {target_id}."
                )
            if normalized_id in sequences:
                raise ValueError(f"duplicate on-target site {normalized_id}.")
            sequences[normalized_id] = sequence
        if candidate_count != source.expected_candidate_site_count:
            raise ValueError(
                "candidate-site count does not match the source profile."
            )
        if len(sequences) != source.expected_on_target_site_count:
            raise ValueError(
                "on-target-site count does not match the source profile."
            )
        return sequences


def _read_observations(
    content: bytes,
    target_sequences: dict[str, str],
    source: CynomolgusBaseEditingSource,
) -> tuple[_ObservedBaseRecord, ...]:
    records: list[_ObservedBaseRecord] = []
    embryo_base_observation_count = 0
    clone_denominator_count = 0
    with _publisher_workbook(content) as workbook:
        sheet_name = "Genotyping by Sanger Sequencing"
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"source-data workbook is missing {sheet_name}.")
        sheet = workbook[sheet_name]
        if not str(sheet.cell(1, 1).value).startswith(
            "Source Data for genotyping by Sanger sequencing"
        ):
            raise ValueError("source-data workbook title is not qualified.")
        for block in source.blocks:
            _validate_block(block, sheet, target_sequences)
            intended_counts = [0] * len(block.target_bases)
            clone_denominator = 0
            embryo_ids: set[int] = set()
            for row_number in range(
                block.first_data_row,
                block.last_data_row + 1,
            ):
                embryo_id = sheet.cell(row_number, 1).value
                total_clones = sheet.cell(row_number, 2).value
                if (
                    isinstance(embryo_id, bool)
                    or not isinstance(embryo_id, int)
                    or embryo_id < 1
                    or embryo_id in embryo_ids
                ):
                    raise ValueError(
                        f"{block.target_site_id} has an invalid embryo ID."
                    )
                if (
                    isinstance(total_clones, bool)
                    or not isinstance(total_clones, int)
                    or total_clones < 1
                ):
                    raise ValueError(
                        f"{block.target_site_id} has an invalid clone count."
                    )
                embryo_ids.add(embryo_id)
                clone_denominator += total_clones
                for index, column in enumerate(
                    block.intended_count_columns
                ):
                    count = _source_count(
                        sheet.cell(row_number, column).value,
                        block.target_site_id,
                    )
                    if count > total_clones:
                        raise ValueError(
                            f"{block.target_site_id} intended count exceeds "
                            "the clone denominator."
                        )
                    intended_counts[index] += count
            for target_base, intended_count in zip(
                block.target_bases,
                intended_counts,
                strict=True,
            ):
                records.append(
                    _ObservedBaseRecord(
                        record_id=(
                            f"{block.context_id}:{block.target_site_id}:"
                            f"{target_base}"
                        ),
                        context_id=block.context_id,
                        editor=block.editor,
                        multiplex_guide_count=block.multiplex_guide_count,
                        target_site_id=block.target_site_id,
                        gene=block.gene,
                        conversion=block.conversion,
                        target_base=target_base,
                        target_sequence=target_sequences[
                            block.target_site_id
                        ],
                        intended_count=intended_count,
                        clone_denominator=clone_denominator,
                        embryo_count=len(embryo_ids),
                    )
                )
            embryo_base_observation_count += (
                len(embryo_ids) * len(block.target_bases)
            )
            clone_denominator_count += (
                clone_denominator * len(block.target_bases)
            )
    if len(records) != source.expected_record_count:
        raise ValueError("record count does not match the source profile.")
    if len({record.context_id for record in records}) != (
        source.expected_context_count
    ):
        raise ValueError("context count does not match the source profile.")
    if embryo_base_observation_count != (
        source.expected_embryo_base_observation_count
    ):
        raise ValueError(
            "embryo-base observation count does not match the source profile."
        )
    if clone_denominator_count != source.expected_clone_denominator_count:
        raise ValueError(
            "clone denominator count does not match the source profile."
        )
    return tuple(records)


@contextmanager
def _publisher_workbook(content: bytes) -> Iterator[Any]:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=(
                "Unknown extension is not supported and will be removed"
            ),
            category=UserWarning,
            module=r"openpyxl\.worksheet\._reader",
        )
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
        )
        try:
            yield workbook
        finally:
            workbook.close()


def _validate_block(
    block: CynomolgusBaseEditingBlock,
    sheet,
    target_sequences: dict[str, str],
) -> None:
    if block.target_site_id not in target_sequences:
        raise ValueError(
            f"source profile target {block.target_site_id} is unavailable."
        )
    if not block.context_id or not block.editor or not block.gene:
        raise ValueError("source block identity fields are required.")
    if block.conversion not in {"C_to_T", "A_to_G"}:
        raise ValueError("source block conversion must be C_to_T or A_to_G.")
    if (
        isinstance(block.multiplex_guide_count, bool)
        or not 1 <= block.multiplex_guide_count <= 3
    ):
        raise ValueError("multiplex_guide_count must be between 1 and 3.")
    if len(block.target_bases) != len(block.intended_count_columns):
        raise ValueError("source block target bases and columns must align.")
    expected_source_label = block.source_label or block.target_site_id
    if sheet.cell(block.label_row, 1).value != expected_source_label:
        raise ValueError(
            f"source-data block label mismatch for {block.target_site_id}."
        )
    if sheet.cell(block.header_row, 1).value != "Embryo ID":
        raise ValueError(
            f"source-data embryo header mismatch for {block.target_site_id}."
        )
    if "Number of Total clone" not in str(
        sheet.cell(block.header_row, 2).value
    ):
        raise ValueError(
            f"source-data clone header mismatch for {block.target_site_id}."
        )
    for target_base, column in zip(
        block.target_bases,
        block.intended_count_columns,
        strict=True,
    ):
        if sheet.cell(block.label_row, column).value != target_base:
            raise ValueError(
                f"source-data target-base mismatch for "
                f"{block.target_site_id}:{target_base}."
            )
        header = str(sheet.cell(block.header_row, column).value)
        expected = "C>T" if block.conversion == "C_to_T" else "A>G"
        if expected not in header:
            raise ValueError(
                f"source-data conversion header mismatch for "
                f"{block.target_site_id}:{target_base}."
            )


def _source_count(value: object, target_site_id: str) -> int:
    match = _COUNT_PATTERN.match(str(value))
    if not match:
        raise ValueError(
            f"{target_site_id} contains an invalid source count {value!r}."
        )
    return int(match.group(1))
