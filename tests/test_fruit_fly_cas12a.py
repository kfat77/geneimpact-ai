from hashlib import sha256
from io import BytesIO
import json
import sys

from openpyxl import Workbook
import pytest

from geneimpact.fruit_fly_cas12a import (
    FruitFlyCas12aSource,
    audit_fruit_fly_cas12a_evidence,
    lookup_fruit_fly_cas12a_array,
)
from geneimpact.cli import main


GUIDES = (
    "ACGTACGTACGTACGTACGTACG",
    "CGTACGTACGTACGTACGTACGT",
    "GTACGTACGTACGTACGTACGTA",
    "TACGTACGTACGTACGTACGTAC",
)


def _workbook_bytes(workbook):
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _synthetic_sources():
    library = (
        "Supplementary Table 1: Synthetic HD12aCFD library,,,,,,\n"
        "Line_ID,sgRNA_1,sgRNA_2,sgRNA_3,sgRNA_4,"
        "primary_gene,primary_gene_symbol\n"
        f"HD12aCFD0001,{','.join(GUIDES)},FBgn0000001,gene1\n"
        f"HD12aCFD0002L,{','.join(GUIDES[:3])},,FBgn0000002,gene2\n"
    ).encode("utf-8")

    genotypes = Workbook()
    genotype_sheet = genotypes.active
    genotype_sheet.title = "Sheet1"
    genotype_sheet.append(
        ["Supplementary Table 3: Genotypes of the Drosophila strains."]
    )
    genotype_sheet.append(
        [
            "Stock_Name",
            "Genotype",
            "Used in",
            "Internal_Transgene_IDs_Source",
            "Comment",
        ]
    )
    genotype_sheet.append(
        [
            "HD12aCFD sgRNA lines",
            "w[*]; P{HD12aCFD}attP40 / CyO-GFP",
            "Fig. 1-7",
            "synthetic",
            None,
        ]
    )
    genotype_bytes = _workbook_bytes(genotypes)

    source_data = Workbook()
    fig5i = source_data.active
    fig5i.title = "Fig5i"
    fig5i.append(["Line_ID", "CIMR_score", "Comment", "Disc_size", "Date"])
    fig5i.append(["emptyCFD8", 0, "NA", 0, None])
    fig5i.append(["HD12aCFD0001", 0, "NA", 0, None])
    fig5i.append(["HD12aCFD0001", 2, "NA", 0, None])
    fig5i.append(["HD12aCFD0002", "NA", "low quality", 0, None])
    fig5j = source_data.create_sheet("Fig5j")
    fig5j.append(
        ["sgRNA_Line", "LOH_0", "LOH_1", "LOH_2", "LOH_3", "comment"]
    )
    fig5j.append(["HD12aCFD001", 1, 2, 0, 1, None])
    fig5j.append(["HD12aCFD002", None, None, None, None, "not scored"])
    source_bytes = _workbook_bytes(source_data)
    return library, genotype_bytes, source_bytes


def _source_profile(library, genotypes, source_data):
    return FruitFlyCas12aSource(
        source_id="synthetic-dmel-cas12a-array-evidence",
        article_reference="https://example.test/cas12a",
        library_url="https://example.test/library.csv",
        genotypes_url="https://example.test/genotypes.xlsx",
        source_data_url="https://example.test/source.xlsx",
        library_sha256=sha256(library).hexdigest(),
        genotypes_sha256=sha256(genotypes).hexdigest(),
        source_data_sha256=sha256(source_data).hexdigest(),
        expected_library_array_count=2,
        expected_library_guide_count=7,
        expected_fig5i_row_count=4,
        expected_fig5i_numeric_array_observation_count=2,
        expected_fig5i_missing_array_observation_count=1,
        expected_fig5i_control_observation_count=1,
        expected_fig5i_unique_array_count=2,
        expected_fig5j_row_count=2,
        expected_fig5j_scored_array_count=1,
        expected_fig5j_disc_count=4,
    )


def _write_sources(tmp_path):
    library, genotypes, source_data = _synthetic_sources()
    library_path = tmp_path / "library.csv"
    genotypes_path = tmp_path / "genotypes.xlsx"
    source_data_path = tmp_path / "source-data.xlsx"
    library_path.write_bytes(library)
    genotypes_path.write_bytes(genotypes)
    source_data_path.write_bytes(source_data)
    return (
        library_path,
        genotypes_path,
        source_data_path,
        _source_profile(library, genotypes, source_data),
    )


def test_audits_pinned_array_level_evidence_without_predictor_promotion(
    tmp_path,
):
    library, genotypes, source_data, source = _write_sources(tmp_path)

    report = audit_fruit_fly_cas12a_evidence(
        library,
        genotypes,
        source_data,
        source=source,
    )

    assert report.species_profile == "fruit_fly"
    assert report.nuclease == "LbCas12a-D156R"
    assert report.reagent_level == "three_or_four_guide_array"
    assert report.library_array_count == 2
    assert report.library_guide_count == 7
    assert report.fig5i_numeric_array_observation_count == 2
    assert report.fig5j_disc_count == 4
    assert report.source_verification == "all_pinned_sources_verified"
    assert report.predictive_adapter_available is False
    assert report.discrimination_metrics_status == (
        "not_applicable_extreme_class_imbalance"
    )
    assert not any(guide in repr(report) for guide in GUIDES)


def test_looks_up_indivisible_array_evidence_with_sequence_hashes(tmp_path):
    library, genotypes, source_data, source = _write_sources(tmp_path)

    evidence = lookup_fruit_fly_cas12a_array(
        library,
        genotypes,
        source_data,
        "HD12aCFD0002",
        source=source,
    )

    assert evidence.source_line_id == "HD12aCFD0002L"
    assert evidence.component_guide_count == 3
    assert evidence.component_sequence_sha256 == tuple(
        sha256(guide.encode("ascii")).hexdigest() for guide in GUIDES[:3]
    )
    assert evidence.fig5i_numeric_observation_count == 0
    assert evidence.fig5i_missing_observation_count == 1
    assert evidence.fig5j_disc_count == 0
    assert evidence.interpretation == "array_level_loh_observation_only"
    assert not any(guide in repr(evidence) for guide in GUIDES)


def test_rejects_unpinned_source_content(tmp_path):
    library, genotypes, source_data, source = _write_sources(tmp_path)
    library.write_bytes(library.read_bytes() + b"\n")

    with pytest.raises(ValueError, match="library SHA-256"):
        audit_fruit_fly_cas12a_evidence(
            library,
            genotypes,
            source_data,
            source=source,
        )


def test_cli_writes_audit_and_optional_array_lookup(
    tmp_path,
    monkeypatch,
    capsys,
):
    library, genotypes, source_data, source = _write_sources(tmp_path)
    output = tmp_path / "fruit-fly-audit.json"
    monkeypatch.setattr(
        "geneimpact.cli.PORT_2026_CAS12A_SOURCE",
        source,
    )
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "geneimpact",
            "audit-fruit-fly-cas12a-evidence",
            "--library",
            str(library),
            "--genotypes",
            str(genotypes),
            "--source-data",
            str(source_data),
            "--line-id",
            "HD12aCFD002",
            "--output",
            str(output),
        ],
    )

    main()

    report = json.loads(output.read_text(encoding="utf-8"))
    assert report["audit"]["predictive_adapter_available"] is False
    assert report["array_evidence"]["source_line_id"] == "HD12aCFD0002L"
    assert "written" in capsys.readouterr().out
