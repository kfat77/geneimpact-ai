from hashlib import sha256
from io import BytesIO
import json
import sys

from openpyxl import Workbook
import pytest

import geneimpact.cli as cli
from geneimpact.cli import main
from geneimpact.rat_validation import (
    RatGuideActivitySource,
    evaluate_rat_guide_transfer,
    prepare_rat_guide_transfer_template,
)


TARGET_TO_GUIDE = (
    ("rat Il13_ON", "rIl13_gRNA"),
    ("rat Map3k14_ON", "rMap3k14_gRNA"),
    ("rat Trpa1_ON", "rTrpa1_1_gRNA"),
    ("rat Usp30_ON", "rUsp30_1_gRNA"),
    ("rEsr1_sgRNA1_ON", "rEsr1_1_gRNA"),
    ("rEsr1_sgRNA2_ON", "rEsr1_2_gRNA"),
    ("rIL33_ON", "rIl33_gRNA"),
    ("rJag1_sgRNA1_ON", "rJag1_1_gRNA"),
    ("rJag1_sgRNA2_ON", "rJag1_2_gRNA"),
    ("rMap4k1_ON", "rMap4k1_gRNA"),
    ("rRipk1_ON", "rRipk1_gRNA"),
    ("rRorc_sgRNA1_ON", "rRorc_1_gRNA"),
    ("rRorc_sgRNA2_ON", "rRorc_2_gRNA"),
    ("rRorc_sgRNA3_ON", "rRorc_3_gRNA"),
)
SEQUENCES = {
    guide: f"{index:020b}".replace("0", "A").replace("1", "C")
    for index, (_, guide) in enumerate(TARGET_TO_GUIDE, start=1)
}
SEQUENCES[TARGET_TO_GUIDE[0][1]] = SEQUENCES[TARGET_TO_GUIDE[0][1]][1:]
OBSERVED = {
    target: 0.20 + index * 0.04
    for index, (target, _) in enumerate(TARGET_TO_GUIDE, start=1)
}
ANIMAL_COUNTS = {
    target: index + 2
    for index, (target, _) in enumerate(TARGET_TO_GUIDE, start=1)
}


def _workbook_bytes(workbook):
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _source_workbooks():
    table1 = Workbook()
    sheet1 = table1.active
    sheet1.title = "gRNA ON_OFF Target List"
    sheet1.append(
        [
            "Name",
            "5'+1nt Target Sequence\n(Actual)",
            "Target Sequence\n(Design)",
            "PAM",
            "Chromosome",
            "Chr Position",
            "% of animals w/OT",
            "Genome",
            "Mismatch #",
            "MIT Score",
            "# of OFF-targets",
        ]
    )
    for _, guide in TARGET_TO_GUIDE:
        sequence = SEQUENCES[guide]
        sheet1.append(
            [
                guide,
                f"G{sequence}",
                sequence,
                None,
                "chr1",
                1,
                "gRNA",
                "rn5",
                None,
                50,
                0,
            ]
        )
    for index in range(11):
        sheet1.append(
            [
                f"rExtra_{index}_gRNA",
                "G" + "T" * 20,
                "T" * 20,
                None,
                "chr2",
                index + 2,
                "gRNA",
                "rn5",
                None,
                50,
                0,
            ]
        )

    table5 = Workbook()
    positive = table5.active
    positive.title = "Suppl.Fig.8a OT-positive"
    negative = table5.create_sheet("Suppl.Fig.8a OT-negative")
    headers = ["Target", "Mean ON-target efficiency", "Animal number (n)"]
    positive.append(headers)
    negative.append(["Target", "Mean ON target efficiency", "Animal number (n)"])
    for index, (target, _) in enumerate(TARGET_TO_GUIDE):
        row = [target, OBSERVED[target], ANIMAL_COUNTS[target]]
        (positive if index < 4 else negative).append(row)
    negative.append(["rUsp30_sgRNA1_ON", 0.7, 9])
    negative.append(["rUsp30_sgRNA2_ON", 0.8, 9])
    return _workbook_bytes(table1), _workbook_bytes(table5)


def _source_profile(table1_bytes, table5_bytes):
    return RatGuideActivitySource(
        source_id="synthetic-rat-transfer-test",
        article_reference="https://example.test/rat-source",
        table1_url="https://example.test/table1.xlsx",
        table5_url="https://example.test/table5.xlsx",
        table1_sha256=sha256(table1_bytes).hexdigest(),
        table5_sha256=sha256(table5_bytes).hexdigest(),
        source_genome_build="rn5",
        expected_rat_guide_count=25,
        expected_label_count=16,
        target_to_guide=TARGET_TO_GUIDE,
        excluded_targets=(
            "rUsp30_sgRNA1_ON",
            "rUsp30_sgRNA2_ON",
        ),
    )


def _predictions():
    return {
        "prediction": {
            "name": "synthetic-oracle",
            "version": "1.0",
            "score_direction": "higher_is_more_active",
            "score_semantics": "probability",
            "sequence_basis": "design_sequence",
            "training_overlap_status": "declared_no_overlap",
            "evidence_reference": "synthetic-independent-record",
        },
        "records": [
            {
                "target": target,
                "design_sequence_sha256": sha256(
                    SEQUENCES[guide].encode("ascii")
                ).hexdigest(),
                "design_sequence_length": len(SEQUENCES[guide]),
                "actual_guide_sequence_sha256": sha256(
                    f"G{SEQUENCES[guide]}".encode("ascii")
                ).hexdigest(),
                "actual_guide_sequence_length": len(SEQUENCES[guide]) + 1,
                "predicted_score": OBSERVED[target],
            }
            for target, guide in TARGET_TO_GUIDE
        ],
    }


def test_evaluates_verified_rat_in_vivo_transfer_without_retaining_sequences(
    tmp_path,
):
    table1_bytes, table5_bytes = _source_workbooks()
    table1_path = tmp_path / "table1.xlsx"
    table5_path = tmp_path / "table5.xlsx"
    table1_path.write_bytes(table1_bytes)
    table5_path.write_bytes(table5_bytes)

    report = evaluate_rat_guide_transfer(
        table1_path,
        table5_path,
        _predictions(),
        source=_source_profile(table1_bytes, table5_bytes),
    )

    assert report.source_verification == "pinned_workbooks_verified"
    assert report.source_table1_url == "https://example.test/table1.xlsx"
    assert "not redistributed" in report.source_license_note
    assert report.species_profile == "rat"
    assert report.guide_count == 14
    assert report.excluded_ambiguous_guide_count == 2
    assert report.metrics.pearson_r == pytest.approx(1.0)
    assert report.metrics.spearman_rho == pytest.approx(1.0)
    assert report.metrics.mean_absolute_error == pytest.approx(0.0)
    assert report.training_overlap_status == "declared_no_overlap"
    assert report.independence_verified is False
    assert "submitter-declared" in report.independence_interpretation
    assert report.use == "external_transfer_ranking_benchmark_only"
    assert not any(sequence in repr(report) for sequence in SEQUENCES.values())


def test_prepares_sequence_redacted_prediction_template(tmp_path):
    table1_bytes, table5_bytes = _source_workbooks()
    table1_path = tmp_path / "table1.xlsx"
    table5_path = tmp_path / "table5.xlsx"
    table1_path.write_bytes(table1_bytes)
    table5_path.write_bytes(table5_bytes)

    template = prepare_rat_guide_transfer_template(
        table1_path,
        table5_path,
        source=_source_profile(table1_bytes, table5_bytes),
    )

    assert template["prediction"]["training_overlap_status"] == "unknown"
    assert len(template["records"]) == 14
    assert template["records"][0]["predicted_score"] is None
    assert template["records"][0]["design_sequence_sha256"] == sha256(
        SEQUENCES[TARGET_TO_GUIDE[0][1]].encode("ascii")
    ).hexdigest()
    assert template["records"][0]["design_sequence_length"] == 19
    assert template["records"][0]["actual_guide_sequence_length"] == 20
    assert not any(
        sequence in repr(template)
        for design in SEQUENCES.values()
        for sequence in (design, f"G{design}")
    )


def test_rat_transfer_cli_writes_auditable_json(
    tmp_path,
    monkeypatch,
    capsys,
):
    table1_bytes, table5_bytes = _source_workbooks()
    table1_path = tmp_path / "table1.xlsx"
    table5_path = tmp_path / "table5.xlsx"
    predictions_path = tmp_path / "predictions.json"
    output_path = tmp_path / "report.json"
    table1_path.write_bytes(table1_bytes)
    table5_path.write_bytes(table5_bytes)
    predictions_path.write_text(json.dumps(_predictions()), encoding="utf-8")
    expected = evaluate_rat_guide_transfer(
        table1_path,
        table5_path,
        _predictions(),
        source=_source_profile(table1_bytes, table5_bytes),
    )
    called = {}

    def fake_evaluate(table1, table5, predictions):
        called["paths"] = (table1, table5)
        called["predictions"] = predictions
        return expected

    monkeypatch.setattr(cli, "evaluate_rat_guide_transfer", fake_evaluate)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "geneimpact",
            "validate-rat-guide-transfer",
            "--table1",
            str(table1_path),
            "--table5",
            str(table5_path),
            "--predictions",
            str(predictions_path),
            "--output",
            str(output_path),
        ],
    )

    main()

    written = json.loads(output_path.read_text(encoding="utf-8"))
    assert called["paths"] == (table1_path, table5_path)
    assert called["predictions"]["prediction"]["name"] == "synthetic-oracle"
    assert written["guide_count"] == 14
    assert "written" in capsys.readouterr().out
